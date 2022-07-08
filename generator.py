import json
import os
import shutil
import sys

from loguru import logger

from openpyxl import load_workbook

logger.remove()
handler_id = logger.add(sys.stderr, level="ERROR")
file_paths = os.listdir("excels")


# shutil.rmtree("results")


def generate_sprites(_count, _positions, _corner, _shadow, _border, _width, _height):
    _sprites = []
    _count = 1
    for position in _positions:
        _sprites.append({
            "id": f"sprite_{_count}",
            "height": _height,
            "width": _width,
            "radii": 15 if _corner == "默认" else None,
            "corner": {
                "enable": True if _corner == "默认" else False
            },
            "border": {
                "enable": True if _border == "默认" else False,
                "color": "black",
                "width": 1
            },
            "shadow": {
                "enable": True if _shadow == "默认" else False,
                "iterations": 3 if _shadow == "默认" else _shadow,
                "border": 3,
                "offset_x": 3,
                "offset_y": 3,
                "shadow_color": "grey"
            },
            "rotate": 0,
            "position": eval(position)
        })
        _count = _count + 1
    return _sprites


for file_path in file_paths:
    all_content = []
    check_content = []
    all_path = f"excels/{file_path}/拼图样式.xlsx"
    wb = load_workbook(all_path)
    sheet = wb['Sheet1']
    res = sheet.cell(1, 1).value  # 取到第一行第一列的值
    for i in range(1, sheet.max_row):
        name = sheet.cell(i + 1, 1).value
        if name is None:
            continue
        count = sheet.cell(i + 1, 2).value
        width = sheet.cell(i + 1, 3).value
        height = sheet.cell(i + 1, 4).value
        try:
            positions = sheet.cell(i + 1, 5).value.split("/")
        except:
            logger.error(f"数据格式错误{all_path}/{sheet.cell(i + 1, 1).value}")
            continue
        corner = sheet.cell(i + 1, 6).value
        shadow = sheet.cell(i + 1, 7).value
        border = sheet.cell(i + 1, 8).value
        total_dir = f"results/{name}"
        data_dir = f"excels/{file_path}/{name}"
        shutil.copytree(data_dir, total_dir)
        sprites = generate_sprites(count, positions, corner, shadow, border, width, height)
        model = {'id': name,
                 "type": "template",
                 "version": 1.0,
                 "description": "tbd",
                 "thumbnail": f"templates/{name}/thumbnail.jpg",
                 "background": {
                     "type": "image",
                     "path": f"templates/{name}/images/background.jpg",
                     "height": 800,
                     "width": 800,
                     "radii": None,
                     "corner": {
                         "enable": True if corner == "默认" or corner != "否" else False
                     },
                     "border": {
                         "enable": True if corner == "默认" or corner != "否" else False
                     },
                     "shadow": {
                         "enable": True if corner == "默认" or corner != "否" else False
                     }
                 },
                 "sprites": {
                     "count": count,
                     "list": sprites
                 }
                 }
        content = json.dumps(model, indent=4, ensure_ascii=False)
        all_content.append(content)
        # with open(f'results/{name}/model.json', 'w', encoding="utf-8") as json_file:
        #     result = json.dumps(content, ensure_ascii=False, indent=4)
        #     json_file.write(content)
        # logger.info(f"生成{name}模板")
    for content in all_content:
        json_content = json.loads(content)
        name = json_content['id']
        with open(f'results/{name}/model.json', 'w', encoding="utf-8") as json_file:
            result = json.dumps(content, ensure_ascii=False, indent=4)
    #         if check_template(json_content):
    #             check_content.append(name)
    #         json_file.write(content)
    #         logger.info(f"生成{name}模板")
    # if len(check_content) > 0:
    #     logger.error(f'坐标问题模板:{check_content}')
